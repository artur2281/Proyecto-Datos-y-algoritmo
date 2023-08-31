from base_de_datos import BaseDeDatos
from persona import Persona
import openpyxl
from openpyxl import Workbook

#clase para el registro de personas
class RegistroPersona:
    #funcion para inicializar el registro
    def __init__(self):
        self.bd = BaseDeDatos('personas.xlsx')
        self.bd.cerrar()
        datos = self.bd.obtener_datos()
        if isinstance(datos, list) and all(isinstance(d, dict) for d in datos):
            self.personas = [Persona(**persona) for persona in datos]
        else:
            self.personas = []

    #funcion para agregar una persona al registro
    def agregar_persona(self, persona):
        try:
            wb = openpyxl.load_workbook('personas.xlsx')
            sheet = wb.active
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active
            sheet.append(['Nombre', 'Codigo', 'Edad', 'Correo', 'Número', 'Género', 'Fecha de Nacimiento'])
        sheet.append([persona.nombre, persona.codigo, persona.edad, persona.correo, persona.numero, persona.genero, persona.fecha_nacimiento])
        wb.save('personas.xlsx')
        self.personas.append(persona)
        print("Persona agregada exitosamente.")

    #funcion para eliminar una persona del registro
    def eliminar_persona(self, codigo):
        personas_filtradas = [persona for persona in self.personas if persona.codigo != codigo]
        if len(personas_filtradas) < len(self.personas):
            self.personas = personas_filtradas
            self.bd.guardar_datos([persona.__dict__ for persona in self.personas])
            print(f"Persona {codigo} eliminada.")
        else:
            print(f"No se encontró a {codigo} en el registro.")

    #funcion para mostrar las personas del registro
    def mostrar_personas(self):
        for persona in self.personas:
            print(f"Nombre: {persona.nombre}\nCodigo: {persona.codigo}\nEdad: {persona.edad}\nCorreo: {persona.correo}\nNúmero: {persona.numero}\nGénero: {persona.genero}\nFecha de Nacimiento: {persona.fecha_nacimiento}\n")

    #funcion para buscar una persona en el registro por codigo
    def buscar_persona_por_codigo(self, codigo):
        personas_encontradas = [persona for persona in self.personas if persona.codigo == codigo]
        if personas_encontradas:
            for persona in personas_encontradas:
                print("Persona encontrada:")
                print(f"Nombre: {persona.nombre}\nCodigo: {persona.codigo}\nEdad: {persona.edad}\nCorreo: {persona.correo}\nNúmero: {persona.numero}\nGénero: {persona.genero}\nFecha de Nacimiento: {persona.fecha_nacimiento}\n")
        else:
            print(f"No se encontró a {codigo} en el registro.")

    #funcion para buscar una persona en el registro por nombre
    def buscar_persona_por_nombre(self, nombre):
        wb = openpyxl.load_workbook('personas.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            if row[0] == nombre:
                print(f"Nombre: {row[0]}\nCodigo: {row[1]}\nEdad: {row[2]}\nCorreo: {row[3]}\nNúmero: {row[4]}\nGénero: {row[5]}\nFecha de Nacimiento: {row[6]}\n")
                break
        else:
            print(f"No se encontró a {nombre} en el registro.")

