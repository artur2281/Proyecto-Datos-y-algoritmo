import openpyxl
from openpyxl import Workbook

class BaseDeDatos:
    def __init__(self, archivo):
        self.archivo = archivo

    def obtener_datos(self):
        try:
            wb = openpyxl.load_workbook(self.archivo)
            sheet = wb.active
            datos = []
            for row in sheet.iter_rows(values_only=True):
                if row[0] != 'Nombre':
                    datos.append({'nombre': row[0], 'codigo': row[1], 'edad': row[2], 'correo': row[3], 'numero': row[4], 'genero': row[5], 'fecha_nacimiento': row[6]})
            return datos
        except FileNotFoundError:
            return []

    def guardar_datos(self, datos):
        wb = Workbook()
        sheet = wb.active
        sheet.append(['Nombre', 'Codigo', 'Edad', 'Correo', 'Número', 'Género', 'Fecha de Nacimiento'])
        for persona in datos:
            sheet.append([persona['nombre'], persona['codigo'], persona['edad'], persona['correo'], persona['numero'], persona['genero'], persona['fecha_nacimiento']])
        wb.save(self.archivo)

    def cerrar(self):
        pass
