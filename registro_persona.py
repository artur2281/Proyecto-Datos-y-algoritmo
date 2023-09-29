from base_de_datos import BaseDeDatos
from persona import Persona
import openpyxl
from openpyxl import Workbook
from prettytable import PrettyTable

#clase para el registro de personas
class RegistroPersona:
    #funcion para inicializar el registro
    def __init__(self):
        self.bd = BaseDeDatos('personas.xlsx')
        self.bd.cerrar()
        datos = self.bd.obtener_datos()
        if isinstance(datos, list) and all(isinstance(d, dict) for d in datos):
            self.personas = [Persona(**persona) for persona in datos]
        else:
            self.personas = []

    #funcion para agregar una persona al registro
    def agregar_persona(self, persona):
        try:
            wb = openpyxl.load_workbook('personas.xlsx')
            sheet = wb.active
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active
            sheet.append(['Nombre', 'Codigo', 'Edad', 'Correo', 'Número', 'Género', 'Fecha de Nacimiento'])
        sheet.append([persona.nombre, persona.codigo, persona.edad, persona.correo, persona.numero, persona.genero, persona.fecha_nacimiento])
        wb.save('personas.xlsx')
        self.personas.append(persona)
        print("Persona agregada exitosamente.")

    def editar_persona(self, codigo, nueva_informacion):
        persona_editada = None
        for persona in self.personas:
            if persona.codigo == codigo:
                persona.nombre = nueva_informacion.get('nombre', persona.nombre)
                persona.edad = nueva_informacion.get('edad', persona.edad)
                persona.correo = nueva_informacion.get('correo', persona.correo)
                persona.numero = nueva_informacion.get('numero', persona.numero)
                persona.genero = nueva_informacion.get('genero', persona.genero)
                persona.fecha_nacimiento = nueva_informacion.get('fecha_nacimiento', persona.fecha_nacimiento)
                persona_editada = persona
                break

        if persona_editada:
            # Guardar los cambios en el archivo Excel
            wb = openpyxl.load_workbook('personas.xlsx')
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                if row[1] == codigo:
                    # Actualizar los valores en la hoja de cálculo
                    row[0] = persona_editada.nombre
                    row[2] = persona_editada.edad
                    row[3] = persona_editada.correo
                    row[4] = persona_editada.numero
                    row[5] = persona_editada.genero
                    row[6] = persona_editada.fecha_nacimiento
                    break
            wb.save('personas.xlsx')
            print(f"Persona {codigo} editada exitosamente.")
        else:
            print(f"No se encontró a {codigo} en el registro.")



    #funcion para eliminar una persona del registro
    def eliminar_persona(self, codigo):
        personas_filtradas = [persona for persona in self.personas if persona.codigo != codigo]
        if len(personas_filtradas) < len(self.personas):
            self.personas = personas_filtradas
            self.bd.guardar_datos([persona.__dict__ for persona in self.personas])
            print(f"Persona {codigo} eliminada.")
        else:
            print(f"No se encontró a {codigo} en el registro.")

    #funcion para mostrar las personas del registro
    def mostrar_personas(self):
        table = PrettyTable()
        table.field_names = ["Nombre", "Codigo", "Edad", "Correo", "Número", "Género", "Fecha de Nacimiento"]
        for persona in self.personas:
            table.add_row([persona.nombre, persona.codigo, persona.edad, persona.correo, persona.numero, persona.genero, persona.fecha_nacimiento])
        print(table)

    #funcion para buscar una persona en el registro por codigo
    def buscar_persona_por_codigo(self, codigo):
        personas_encontradas = [persona for persona in self.personas if persona.codigo == codigo]
        if personas_encontradas:
            for persona in personas_encontradas:
                print("Persona encontrada:")
                print(f"Nombre: {persona.nombre}\nCodigo: {persona.codigo}\nEdad: {persona.edad}\nCorreo: {persona.correo}\nNúmero: {persona.numero}\nGénero: {persona.genero}\nFecha de Nacimiento: {persona.fecha_nacimiento}\n")
        else:
            print(f"No se encontró a {codigo} en el registro.")

    #funcion para buscar una persona en el registro por nombre
    def buscar_persona_por_nombre(self, nombre):
        wb = openpyxl.load_workbook('personas.xlsx')
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            if row[0] == nombre:
                print(f"Nombre: {row[0]}\nCodigo: {row[1]}\nEdad: {row[2]}\nCorreo: {row[3]}\nNúmero: {row[4]}\nGénero: {row[5]}\nFecha de Nacimiento: {row[6]}\n")
                break
        else:
            print(f"No se encontró a {nombre} en el registro.")

