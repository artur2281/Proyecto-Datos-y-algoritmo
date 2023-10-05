from modules.base_de_datos import BaseDeDatos
from modules.persona import Persona
import openpyxl
from openpyxl import Workbook
#from prettytable import PrettyTable

#clase para el registro de personas
class RegistroPersona:
    #funcion para inicializar el registro
    def __init__(self):
        self.bd = BaseDeDatos()
        self.bd.cerrar()
        datos = self.bd.obtener_datos()
        if isinstance(datos, list) and all(isinstance(d, dict) for d in datos):
            self.personas = [Persona(**persona) for persona in datos]
        else:
            self.personas = []

    #funcion para agregar una persona al registro
    def agregar_persona(self, persona):
        try:
            wb = openpyxl.load_workbook('G:\Proyecto final\Proyecto-Datos-y-algoritmo\intefaz\personas.xlsx')
            sheet = wb.active
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active
            sheet.append(['Nombre', 'Codigo', 'Edad', 'Correo', 'Número', 'Género', 'Fecha de Nacimiento'])
        sheet.append([persona.nombre, persona.codigo, persona.edad, persona.correo, persona.numero, persona.genero, persona.fecha_nacimiento])
        wb.save('personas.xlsx')
        self.personas.append(persona)
        print("Persona agregada exitosamente.")

    #funcion para eliminar una persona del registro
    def eliminar_persona(self, codigo):
        personas_filtradas = [persona for persona in self.personas if persona.codigo != codigo]
        if len(personas_filtradas) < len(self.personas):
            self.personas = personas_filtradas
            self.bd.guardar_datos([persona.__dict__ for persona in self.personas])
            print(f"Persona {codigo} eliminada.")
        else:
            print(f"No se encontró a {codigo} en el registro.")



        #funcion para buscar una persona en el registro por codigo
    def editar_persona(self, codigo, nueva_informacion):
        for persona in self.personas:
            if persona.codigo == codigo:
                # Actualizar los datos de la persona con los nuevos valores
                persona.nombre = nueva_informacion.get('nombre', persona.nombre)
                persona.edad = nueva_informacion.get('edad', persona.edad)
                persona.correo = nueva_informacion.get('correo', persona.correo)
                persona.numero = nueva_informacion.get('numero', persona.numero)
                persona.genero = nueva_informacion.get('genero', persona.genero)
                persona.fecha_nacimiento = nueva_informacion.get('fecha_nacimiento', persona.fecha_nacimiento)
                print(f"Datos de {persona.nombre} actualizados exitosamente.")
                return
        print(f"No se encontró a la persona con código {codigo} en el registro.")

#funcion para buscar una persona en el registro por codigo



    #Funcion para extraer el email receptor de la base de datos todos
    def extraer_emails(self):
        wb = openpyxl.load_workbook('G:\Proyecto final\Proyecto-Datos-y-algoritmo\intefaz\personas.xlsx')
        
        sheet = wb.active
        emails_receptores = []  # Lista para almacenar los correos electrónicos
        for row in sheet.iter_rows(values_only=True):
            email_receptor = row[3]
            emails_receptores.append(email_receptor)
        return emails_receptores
    
    #funcion para extraer el email receptor segun la fecha del ordenador
    