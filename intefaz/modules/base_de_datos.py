import openpyxl
from openpyxl import Workbook
from persona import Persona


class BaseDeDatos:
    def __init__(self):

        #inicialisamos donde se encuentre el archivo
        self.archivo = 'G:\Proyecto final\Proyecto-Datos-y-algoritmo\intefaz\personas.xlsx'
        datos = self.obtener_datos()
        if isinstance(datos, list) and all(isinstance(d, dict) for d in datos):
            self.personas = [Persona(**persona) for persona in datos]
        else:
            self.personas = []

    def obtener_datos(self):
        try:
            wb = openpyxl.load_workbook(self.archivo)
            sheet = wb.active
            datos = []
            for row in sheet.iter_rows(values_only=True):
                if row[0] != 'Nombre':
                    datos.append({'nombre': row[0], 'codigo': row[1], 'edad': row[2], 'correo': row[3], 'numero': row[4], 'genero': row[5], 'fecha_nacimiento': row[6]})
            return datos
        except FileNotFoundError:
            return []

    def guardar_datos(self, datos):
        wb = Workbook()
        sheet = wb.active
        sheet.append(['Nombre', 'Codigo', 'Edad', 'Correo', 'Número', 'Género', 'Fecha de Nacimiento'])
        for persona in datos:
            sheet.append([persona['nombre'], persona['codigo'], persona['edad'], persona['correo'], persona['numero'], persona['genero'], persona['fecha_nacimiento']])
        wb.save(self.archivo)
        #funcion para mostrar las personas del registro
    def mostrar_personas(self):
        tabless = []
        for persona in self.personas:
            tabless.append([persona.nombre, persona.codigo, persona.edad, persona.correo, persona.numero, persona.genero, persona.fecha_nacimiento])

        tabla =  tabless 
        return tabla, len(self.personas)
    

    def buscar_persona_por_codigo(self, codigo):
        #busca ña érspma por le codigo 
        personas_encontradas = [persona for persona in self.personas if persona.codigo == codigo]
        if personas_encontradas:
            for persona in personas_encontradas:
                print("Persona encontrada:")
                print(f"Nombre: {persona.nombre}\nCodigo: {persona.codigo}\nEdad: {persona.edad}\nCorreo: {persona.correo}\nNúmero: {persona.numero}\nGénero: {persona.genero}\nFecha de Nacimiento: {persona.fecha_nacimiento}\n")
                persona_encontras = [persona.nombre, persona.codigo,persona.edad,persona.correo, persona.numero,persona.genero, persona.fecha_nacimiento]
                persona_encontrada = persona_encontras
                return persona_encontrada, len(personas_encontradas)
        else:
            print(f"No se encontró a {codigo} en el registro.")


        #funcion para eliminar una persona del registro
    def eliminar_persona(self, codigo):
        personas_filtradas = [persona for persona in self.personas if persona.codigo != codigo]
        if len(personas_filtradas) < len(self.personas):
            self.personas = personas_filtradas
            self.guardar_datos([persona.__dict__ for persona in self.personas])
            print(f"Persona {codigo} eliminada.")
        else:
            print(f"No se encontró a {codigo} en el registro.")

    def agregar_persona(self, persona):
        try:
            wb = openpyxl.load_workbook(self.archivo)
            sheet = wb.active
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active
            sheet.append(['Nombre', 'Codigo', 'Edad', 'Correo', 'Número', 'Género', 'Fecha de Nacimiento'])
        sheet.append([persona.nombre, persona.codigo, persona.edad, persona.correo, persona.numero, persona.genero, persona.fecha_nacimiento])
        wb.save('personas.xlsx')
        self.personas.append(persona)
        print("Persona agregada exitosamente.")


    def editar_persona(self, codigo, nueva_informacion):
        for persona in self.personas:
            if persona.codigo == codigo:
                # Actualizar los datos de la persona con los nuevos valores
                persona.nombre = nueva_informacion.get('nombre', persona.nombre)
                persona.edad = nueva_informacion.get('edad', persona.edad)
                persona.correo = nueva_informacion.get('correo', persona.correo)
                persona.numero = nueva_informacion.get('numero', persona.numero)
                persona.genero = nueva_informacion.get('genero', persona.genero)
                persona.fecha_nacimiento = nueva_informacion.get('fecha_nacimiento', persona.fecha_nacimiento)
                print(f"Datos de {persona.nombre} actualizados exitosamente.")
                return
        print(f"No se encontró a la persona con código {codigo} en el registro.")

#funcion para buscar 
    def editar_persona_por_codigo(self, codigo, nueva_informacion):
        # Buscar la persona por código
        persona_encontrada = None
        for persona in self.personas:
            if persona.codigo == codigo:
                persona_encontrada = persona
                break

        if persona_encontrada:
            # Desempaquetar la nueva información
            nombre = nueva_informacion.get('nombre', persona_encontrada.nombre)
            edad = nueva_informacion.get('edad', persona_encontrada.edad)
            correo = nueva_informacion.get('correo', persona_encontrada.correo)
            numero = nueva_informacion.get('numero', persona_encontrada.numero)
            genero = nueva_informacion.get('genero', persona_encontrada.genero)
            fecha_nacimiento = nueva_informacion.get('fecha_nacimiento', persona_encontrada.fecha_nacimiento)

            # Actualizar los datos de la persona
            if nombre:
                persona_encontrada.set_nombre(nombre)
            if edad:
                persona_encontrada.set_edad(edad)
            if numero:
                persona_encontrada.set_numero(numero)
            if genero:
                persona_encontrada.set_genero(genero)
            if fecha_nacimiento:
                persona_encontrada.set_fecha_nacimiento(fecha_nacimiento)
            if correo:
                persona_encontrada.set_correo(correo)

            print("Datos de la persona actualizados exitosamente.")
        else:
            print(f"No se encontró a {codigo} en el registro.")


    #funcion no se utiliza 
    def extraer_emails(self):
        wb = openpyxl.load_workbook(self.archivo)
        
        sheet = wb.active
        emails_receptores = []  # Lista para almacenar los correos electrónicos
        for row in sheet.iter_rows(values_only=True):
            email_receptor = row[3]
            emails_receptores.append(email_receptor)
        return emails_receptores

    def cerrar(self):
        pass
